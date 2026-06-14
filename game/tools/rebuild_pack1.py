#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rebuild the first booster pack (garage_pack) economy in startx_data.xlsx.

- Clears the card slots of ALL packs (keeps id/name/stage/price/min/max).
- Refills garage_pack with the new 12-card tutorial economy.
- Adds the new cards (fresh p1_* ids to avoid colliding with the 80 legacy
  recipes; only `founder` and `cash` are reused, neither of which appears as
  an all-pack1 input in any legacy recipe).
- Adds the 7 recipes that wire the two production chains + the cash payout.
"""
import openpyxl

XLSX = "data/startx_data.xlsx"

# ---------------------------------------------------------------- new cards
# id, name, type, workTags, salary, capacity, value, maxUses, workRequired, cost
NEW_CARDS = [
    ("p1_neighborhood", "居住小区", "resource", "", "", "", "", "3-5", "", ""),
    # 小区青年是「客户/被招募对象」而非员工——不能去给节点打工（否则会自己刷自己、还能去批发市场刷产品）
    ("p1_youth",        "小区青年", "customer",      "", "", "", 1, "", 10, ""),
    ("p1_survey",       "市场调研", "tool",      "", "", "", 1, "", "", ""),
    ("p1_customer",     "靠谱客户", "customer",      "", "", "", 2, "", 15, ""),
    ("p1_university",   "大学",     "facility",      "", "", "", "", "", "", ""),
    ("p1_intern",       "实习生",   "employee",      "any", 1, 1, "", "", 8, ""),
    ("p1_wholesale",    "批发市场", "resource", "", "", "", "", "3-5", "", ""),
    ("p1_rawprod",      "裸奔粗糙产品", "product",   "", "", "", 1, "", 15, 1),
    ("p1_marketing",    "营销策划", "tool",      "", "", "", 1, "", "", ""),
    ("p1_package",      "带包装粗糙产品", "product",  "", "", "", 2, "", 30, 2),
    ("p1_office",       "创始人办公桌", "facility",   "", 2, "", "", "", "", ""),
    ("p2_law_firm",     "律师事务所", "resource", "", "", "", "", "3-5", "", ""),
    ("p2_document",     "文书", "tool", "", "", "", 0, "", 8, ""),
    ("p2_contract",     "合同", "tool", "", "", "", 0, "", "", ""),
    ("p2_grad",         "毕业生", "employee", "any", 1, 2, "", "", 6, ""),
    ("p2_admin_management", "行政管理", "tool", "", "", "", 1, "", "", ""),
    ("p2_sales_course", "销售技巧课程", "tool", "", "", "", 2, "", "", ""),
    ("p2_product_course", "产品技能课程", "tool", "", "", "", 2, "", "", ""),
    ("p2_legal_admin_course", "法务行政课程", "tool", "", "", "", 2, "", "", ""),
    ("p2_sales_specialist", "销售专员", "employee", "sales,any", 2, 3, "", "", 6, ""),
    ("p2_product_specialist", "产品专员", "employee", "dev,any", 2, 3, "", "", 6, ""),
    ("p2_admin_specialist", "行政专员", "employee", "admin,any", 2, 3, "", "", 6, ""),
    ("p2_orderly_workstation", "井井有条的工位", "facility", "", "", "", 1, "", "", ""),
]

# --------------------------------------------------------------- new recipes
# Each tuple retains the former duration position for compatibility with this
# maintenance script, but production time is derived only from output workRequired.
NEW_RECIPES = [
    ("p2_specialist_make_customer", "拓展靠谱客户", "", 8,
     [("p2_sales_specialist", 1, False), ("p1_neighborhood", 1, False)], [("p1_customer", 1)]),
    ("p1_recruit_youth", "招募小区青年", "any", 10,
     [("p1_neighborhood", 1, False)], [("p1_youth", 1)]),
    ("p1_make_customer", "转化靠谱客户", "", 15,
     [("p1_survey", 1, True), ("p1_youth", 1, True)], [("p1_customer", 1)]),
    ("p1_recruit_intern", "招募实习生", "any", 8,
     [("p1_university", 1, False)], [("p1_intern", 1)]),
    ("p2_specialist_package_product", "包装粗糙产品", "", 12,
     [("p2_product_specialist", 1, False), ("p1_wholesale", 1, False)], [("p1_package", 1)]),
    ("p1_make_rawprod", "做出粗糙产品", "any", 15,
     [("p1_wholesale", 1, False)], [("p1_rawprod", 1)]),
    ("p1_package_prod", "包装产品", "", 30,
     [("p1_rawprod", 1, True), ("p1_marketing", 1, True)], [("p1_package", 1)]),
    # 任意产品 + 任意客户 = 现金（数量 = ceil((产品cost + 客户value) * 1.5)），现金依次快速跳出
    # 产品: 裸奔粗糙产品(1) / 带包装粗糙产品(2)；客户: 小区青年(1) / 靠谱客户(2)
    ("p1_cash_pkg_cust", "成交", "", 3,
     [("p1_package", 1, True), ("p1_customer", 1, True)], [("cash", 6)]),   # ceil((2+2)*1.5)
    ("p1_cash_pkg_youth", "成交", "", 3,
     [("p1_package", 1, True), ("p1_youth", 1, True)], [("cash", 5)]),      # ceil((2+1)*1.5)
    ("p1_cash_raw_cust", "成交", "", 3,
     [("p1_rawprod", 1, True), ("p1_customer", 1, True)], [("cash", 5)]),   # ceil((1+2)*1.5)
    ("p1_cash_raw_youth", "成交", "", 3,
     [("p1_rawprod", 1, True), ("p1_youth", 1, True)], [("cash", 3)]),      # ceil((1+1)*1.5)
    ("p2_make_document", "起草文书", "any", 8,
     [("p1_office", 1, False)], [("p2_document", 1)]),
    ("p2_make_contract", "起草合同", "", 8,
     [("p2_document", 1, True), ("p2_law_firm", 1, False)], [("p2_contract", 1)]),
    ("p2_train_grad", "毕业转正", "", 6,
     [("p2_contract", 1, True), ("p1_intern", 1, True)], [("p2_grad", 1)]),
    ("p2_make_sales_course", "编写销售技巧课程", "", 5,
     [("p2_document", 1, True), ("p1_survey", 1, True)], [("p2_sales_course", 1)]),
    ("p2_make_product_course", "编写产品技能课程", "", 5,
     [("p2_document", 1, True), ("p1_marketing", 1, True)], [("p2_product_course", 1)]),
    ("p2_make_legal_admin_course", "编写法务行政课程", "", 5,
     [("p2_document", 1, True), ("p2_admin_management", 1, True)], [("p2_legal_admin_course", 1)]),
    ("p2_train_sales_specialist", "培养销售专员", "", 6,
     [("p2_grad", 1, True), ("p2_sales_course", 1, True)], [("p2_sales_specialist", 1)]),
    ("p2_train_product_specialist", "培养产品专员", "", 6,
     [("p2_grad", 1, True), ("p2_product_course", 1, True)], [("p2_product_specialist", 1)]),
    ("p2_train_admin_specialist", "培养行政专员", "", 6,
     [("p2_grad", 1, True), ("p2_legal_admin_course", 1, True)], [("p2_admin_specialist", 1)]),
    ("p2_build_orderly_workstation", "整理办公工位", "", 8,
     [("p1_office", 1, False), ("p2_admin_management", 1, True)], [("p2_orderly_workstation", 1)]),
]

# ----------------------------------------------- garage_pack slot definition
# four slots, each offers a thematic subset (weights are relative)
GARAGE_SLOTS = [
    [("p1_neighborhood", 40), ("p1_university", 60), ("p1_youth", 25)],
    [("p1_survey", 40), ("p1_customer", 35), ("p1_intern", 25)],
    [("p1_wholesale", 40), ("p1_marketing", 35), ("p1_rawprod", 25)],
    [("p1_package", 34), ("p1_office", 33), ("founder", 33)],
]

DEVELOPMENT_SLOTS = [
    [("p2_law_firm", 55), ("p2_contract", 45)],
    [("p1_intern", 25), ("p2_grad", 20), ("p2_admin_management", 30), ("p2_orderly_workstation", 25)],
    [("p1_survey", 34), ("p1_marketing", 33), ("p2_document", 33)],
    [("p2_sales_course", 34), ("p2_product_course", 33), ("p2_legal_admin_course", 33)],
    [("p2_sales_specialist", 34), ("p2_product_specialist", 33), ("p2_admin_specialist", 33)],
]


def col(ws, name):
    H = [c.value for c in ws[1]]
    if name not in H:
        ws.cell(row=1, column=ws.max_column + 1).value = name
        H = [c.value for c in ws[1]]
    return H.index(name) + 1


def set_row(ws, r, name, val):
    ws.cell(row=r, column=col(ws, name)).value = val


def find_row(ws, card_id):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == card_id:
            return r
    return None


def main():
    wb = openpyxl.load_workbook(XLSX)

    # ---- cards ----
    cw = wb["cards"]
    # founder: bump capacity to 5, salary 0
    fr = find_row(cw, "founder")
    set_row(cw, fr, "capacity", 5)
    set_row(cw, fr, "salary", 0)
    cash_row = find_row(cw, "cash")
    if cash_row:
        set_row(cw, cash_row, "workRequired", 3)
    # add / overwrite new cards
    for (cid, name, typ, tags, sal, cap, value, uses, wreq, cost) in NEW_CARDS:
        r = find_row(cw, cid) or (cw.max_row + 1)
        set_row(cw, r, "id", cid)
        set_row(cw, r, "name", name)
        set_row(cw, r, "type", typ)
        set_row(cw, r, "workTags", tags or None)
        set_row(cw, r, "salary", sal if sal != "" else None)
        set_row(cw, r, "capacity", cap if cap != "" else None)
        set_row(cw, r, "spaceCapacity", 30 if cid == "p1_office" else (5 if cid == "p2_orderly_workstation" else None))
        set_row(cw, r, "value", value if value != "" else None)
        set_row(cw, r, "maxUses", uses if uses != "" else None)
        set_row(cw, r, "workRequired", wreq if wreq != "" else None)
        set_row(cw, r, "cost", cost if cost != "" else None)

    # ---- recipes ----
    rw = wb["recipes"]
    duration_col = col(rw, "duration") if "duration" in [c.value for c in rw[1]] else None
    if duration_col:
        rw.delete_cols(duration_col)
    card_names = {
        cw.cell(row=r, column=col(cw, "id")).value:
        cw.cell(row=r, column=col(cw, "name")).value
        for r in range(2, cw.max_row + 1)
    }
    # clear obsolete/renamed recipe rows (blank the whole row so it's skipped)
    OBSOLETE = {"p1_cash_package", "p1_cash_raw"}
    for r in range(2, rw.max_row + 1):
        if rw.cell(row=r, column=1).value in OBSOLETE:
            for cc in range(1, rw.max_column + 1):
                rw.cell(row=r, column=cc).value = None
    for (rid, name, wtags, dur, inputs, outputs) in NEW_RECIPES:
        r = find_row(rw, rid) or (rw.max_row + 1)
        set_row(rw, r, "id", rid)
        set_row(rw, r, "name", name)
        set_row(rw, r, "备注", "+".join(card_names.get(cid, cid) for cid, _cnt, _cons in inputs))
        set_row(rw, r, "requiredIdeaId", None)
        set_row(rw, r, "packId", "garage_pack" if rid.startswith("p1_") else "Developemnt_pack")
        set_row(rw, r, "worker_tags", wtags or None)
        for i in range(5):
            if i < len(inputs):
                cid, cnt, cons = inputs[i]
                set_row(rw, r, "input%d" % (i + 1), cid)
                set_row(rw, r, "input%dCount" % (i + 1), cnt)
                set_row(rw, r, "input%dConsume" % (i + 1), str(cons).lower())
            else:
                set_row(rw, r, "input%d" % (i + 1), None)
                set_row(rw, r, "input%dCount" % (i + 1), None)
                set_row(rw, r, "input%dConsume" % (i + 1), None)
        for i in range(5):
            if i < len(outputs):
                cid, cnt = outputs[i]
                set_row(rw, r, "output%d" % (i + 1), cid)
                set_row(rw, r, "output%dCount" % (i + 1), cnt)
            else:
                set_row(rw, r, "output%d" % (i + 1), None)
                set_row(rw, r, "output%dCount" % (i + 1), None)
        set_row(rw, r, "output_zone", None)

    # ---- packs: clear every pack's slots, then refill garage_pack ----
    pw = wb["packs"]
    slot_cols = [c.value for c in pw[1] if c.value and c.value.startswith("slot")]
    for r in range(2, pw.max_row + 1):
        if pw.cell(row=r, column=1).value is None:
            continue
        for sc in slot_cols:
            pw.cell(row=r, column=col(pw, sc)).value = None
    gr = find_row(pw, "garage_pack")
    for si, cands in enumerate(GARAGE_SLOTS, start=1):
        for ci, (cid, prob) in enumerate(cands, start=1):
            set_row(pw, gr, "slot%dCard%d" % (si, ci), cid)
            set_row(pw, gr, "slot%dProb%d" % (si, ci), prob)

    # ---- refill Developemnt_pack with the second progression pack ----
    dr = find_row(pw, "Developemnt_pack")
    if dr:
        set_row(pw, dr, "price", 5)
        set_row(pw, dr, "minCards", 3)
        set_row(pw, dr, "maxCards", 5)
        for si, cands in enumerate(DEVELOPMENT_SLOTS, start=1):
            for ci, (cid, prob) in enumerate(cands, start=1):
                set_row(pw, dr, "slot%dCard%d" % (si, ci), cid)
                set_row(pw, dr, "slot%dProb%d" % (si, ci), prob)

    wb.save(XLSX)
    print("rebuilt:", XLSX)


if __name__ == "__main__":
    main()
